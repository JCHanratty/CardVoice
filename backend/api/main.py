"""
CardVoice API - FastAPI backend
WebSocket for real-time voice streaming, REST for collection management.
"""
from fastapi import FastAPI, WebSocket, WebSocketDisconnect, Depends, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy.orm import Session
from pydantic import BaseModel
from typing import List, Optional
import json
import asyncio
import sys
import logging

logger = logging.getLogger(__name__)

from db.models import init_db, get_db, backup_db, SessionLocal, CardSet, Card, VoiceSession, DB_PATH, DB_DIR
from voice.engine import VoiceEngine, parse_spoken_numbers, parse_card_quantities, count_cards, format_output

app = FastAPI(title="CardVoice API", version="0.1.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Initialize database on startup
@app.on_event("startup")
def startup():
    init_db()

# Global voice engine instance
voice_engine = None


# ============================================================
# Pydantic Models
# ============================================================

class SetCreate(BaseModel):
    name: str
    year: Optional[int] = None
    brand: Optional[str] = None
    sport: str = "Baseball"

class CardCreate(BaseModel):
    card_number: str
    player: str
    team: str = ""
    rc_sp: str = ""
    insert_type: str = "Base"
    parallel: str = ""
    qty: int = 0

class BulkCardCreate(BaseModel):
    cards: List[CardCreate]

class QtyUpdate(BaseModel):
    card_number: str
    insert_type: str = "Base"
    parallel: str = ""
    qty: int

class BulkQtyUpdate(BaseModel):
    updates: List[QtyUpdate]

class VoiceParseRequest(BaseModel):
    text: str
    set_id: Optional[int] = None
    insert_type: str = "Base"


# ============================================================
# Set Endpoints
# ============================================================

@app.get("/api/sets")
def list_sets(db: Session = Depends(get_db)):
    """List all card sets."""
    sets = db.query(CardSet).order_by(CardSet.year.desc(), CardSet.name).all()
    return [{"id": s.id, "name": s.name, "year": s.year, "brand": s.brand,
             "sport": s.sport, "total_cards": s.total_cards} for s in sets]

@app.post("/api/sets")
def create_set(data: SetCreate, db: Session = Depends(get_db)):
    """Create a new card set."""
    existing = db.query(CardSet).filter(CardSet.name == data.name).first()
    if existing:
        raise HTTPException(400, "Set already exists")
    card_set = CardSet(name=data.name, year=data.year, brand=data.brand, sport=data.sport)
    db.add(card_set)
    db.commit()
    db.refresh(card_set)
    return {"id": card_set.id, "name": card_set.name}

@app.get("/api/sets/{set_id}")
def get_set(set_id: int, db: Session = Depends(get_db)):
    """Get set details with all cards."""
    card_set = db.query(CardSet).filter(CardSet.id == set_id).first()
    if not card_set:
        raise HTTPException(404, "Set not found")
    
    cards = db.query(Card).filter(Card.set_id == set_id).order_by(Card.card_number).all()
    
    return {
        "id": card_set.id,
        "name": card_set.name,
        "year": card_set.year,
        "brand": card_set.brand,
        "total_cards": len(cards),
        "cards": [{
            "id": c.id,
            "card_number": c.card_number,
            "player": c.player,
            "team": c.team,
            "rc_sp": c.rc_sp,
            "insert_type": c.insert_type,
            "parallel": c.parallel,
            "qty": c.qty,
        } for c in cards]
    }

@app.delete("/api/sets/{set_id}")
def delete_set(set_id: int, db: Session = Depends(get_db)):
    """Delete a set and all its cards."""
    card_set = db.query(CardSet).filter(CardSet.id == set_id).first()
    if not card_set:
        raise HTTPException(404, "Set not found")
    db.delete(card_set)
    db.commit()
    return {"deleted": True}


# ============================================================
# Card Endpoints
# ============================================================

@app.post("/api/sets/{set_id}/cards")
def add_cards(set_id: int, data: BulkCardCreate, db: Session = Depends(get_db)):
    """Add cards to a set (bulk)."""
    card_set = db.query(CardSet).filter(CardSet.id == set_id).first()
    if not card_set:
        raise HTTPException(404, "Set not found")
    
    added = 0
    for c in data.cards:
        existing = db.query(Card).filter(
            Card.set_id == set_id,
            Card.card_number == c.card_number,
            Card.insert_type == c.insert_type,
            Card.parallel == c.parallel,
        ).first()
        
        if existing:
            # Update existing - ADD to quantity
            if c.player:
                existing.player = c.player
            if c.team:
                existing.team = c.team
            if c.rc_sp:
                existing.rc_sp = c.rc_sp
            if c.qty > 0:
                existing.qty += c.qty
        else:
            card = Card(
                set_id=set_id,
                card_number=c.card_number,
                player=c.player,
                team=c.team,
                rc_sp=c.rc_sp,
                insert_type=c.insert_type,
                parallel=c.parallel,
                qty=c.qty,
            )
            db.add(card)
            added += 1
    
    card_set.total_cards = db.query(Card).filter(Card.set_id == set_id).count()
    backup_db()
    db.commit()
    return {"added": added, "total": card_set.total_cards}


@app.delete("/api/cards/{card_id}")
def delete_card(card_id: int, db: Session = Depends(get_db)):
    """Delete a single card."""
    card = db.query(Card).filter(Card.id == card_id).first()
    if not card:
        raise HTTPException(404, "Card not found")
    set_id = card.set_id
    db.delete(card)
    db.flush()  # Flush to mark card as deleted
    card_set = db.query(CardSet).filter(CardSet.id == set_id).first()
    if card_set:
        card_set.total_cards = db.query(Card).filter(Card.set_id == set_id).count()
    db.commit()
    return {"deleted": True}


class CardUpdate(BaseModel):
    card_number: Optional[str] = None
    player: Optional[str] = None
    team: Optional[str] = None
    rc_sp: Optional[str] = None
    insert_type: Optional[str] = None
    parallel: Optional[str] = None
    qty: Optional[int] = None


@app.put("/api/cards/{card_id}")
def update_card(card_id: int, data: CardUpdate, db: Session = Depends(get_db)):
    """Update a single card's fields."""
    card = db.query(Card).filter(Card.id == card_id).first()
    if not card:
        raise HTTPException(404, "Card not found")
    if data.card_number is not None:
        card.card_number = data.card_number
    if data.player is not None:
        card.player = data.player
    if data.team is not None:
        card.team = data.team
    if data.rc_sp is not None:
        card.rc_sp = data.rc_sp
    if data.insert_type is not None:
        card.insert_type = data.insert_type
    if data.parallel is not None:
        card.parallel = data.parallel
    if data.qty is not None:
        card.qty = data.qty
    db.commit()
    db.refresh(card)
    return {
        "id": card.id, "card_number": card.card_number, "player": card.player,
        "team": card.team, "rc_sp": card.rc_sp, "insert_type": card.insert_type,
        "parallel": card.parallel, "qty": card.qty,
    }


@app.put("/api/sets/{set_id}/qty")
def update_quantities(set_id: int, data: BulkQtyUpdate, db: Session = Depends(get_db)):
    """Update card quantities (from voice entry or manual)."""
    card_set = db.query(CardSet).filter(CardSet.id == set_id).first()
    if not card_set:
        raise HTTPException(404, "Set not found")
    
    updated = 0
    for u in data.updates:
        card = db.query(Card).filter(
            Card.set_id == set_id,
            Card.card_number == u.card_number,
            Card.insert_type == u.insert_type,
            Card.parallel == u.parallel,
        ).first()
        
        if card:
            card.qty = u.qty
            updated += 1
    
    db.commit()
    return {"updated": updated}


@app.put("/api/sets/{set_id}/voice-qty")
def voice_update_quantities(set_id: int, data: VoiceParseRequest, db: Session = Depends(get_db)):
    """
    Parse voice text and update quantities.
    Each mention of a number = +1 to that card's qty.
    """
    logger.info("===== ENDPOINT CALLED =====")
    card_set = db.query(CardSet).filter(CardSet.id == set_id).first()
    if not card_set:
        raise HTTPException(404, "Set not found")
    
    text = data.text or ""

    # If the user explicitly said 'card', try to parse card-id / qty pairs
    parsed_pairs = []
    updated = 0
    not_found = []

    logger.info(f"[voice_update_quantities] text='{text}', contains 'card'={'card' in text.lower()}")
    
    if 'card' in text.lower():
        pairs = parse_card_quantities(text)
        logger.info(f"[voice_update_quantities] parse_card_quantities returned: {pairs}")
        parsed_pairs = [{'card': p[0], 'qty': p[1], 'confidence': p[2]} for p in pairs]
        logger.info(f"[voice_update_quantities] parsed_pairs: {parsed_pairs}")

        for card_id, qty, conf in pairs:
            card = db.query(Card).filter(
                Card.set_id == set_id,
                Card.card_number == str(card_id),
                Card.insert_type == data.insert_type,
            ).first()

            if card:
                # Auto-apply parsed quantity (set exact quantity)
                card.qty = qty
                updated += 1
            else:
                not_found.append(card_id)

        db.commit()

        return {
            "parsed_pairs": parsed_pairs,
            "updated": updated,
            "not_found": not_found,
        }

    # Fallback: previous behavior (each mention = +1)
    numbers = parse_spoken_numbers(text)
    counts = count_cards(numbers)

    not_found = []
    updated = 0

    for card_num, qty in counts.items():
        card = db.query(Card).filter(
            Card.set_id == set_id,
            Card.card_number == str(card_num),
            Card.insert_type == data.insert_type,
        ).first()

        if card:
            card.qty = qty
            updated += 1
        else:
            not_found.append(card_num)

    db.commit()

    return {
        "parsed_numbers": numbers,
        "counts": counts,
        "updated": updated,
        "not_found": not_found,
        "output": format_output(numbers),
    }


# ============================================================
# Voice Parse Endpoint (no DB, just text → numbers)
# ============================================================

@app.post("/api/voice/parse")
def parse_voice_text(data: VoiceParseRequest):
    """Parse voice text into card numbers. Stateless utility endpoint."""
    numbers = parse_spoken_numbers(data.text)
    return {
        "numbers": numbers,
        "counts": count_cards(numbers),
        "output": format_output(numbers),
        "unique": len(set(numbers)),
        "total": len(numbers),
    }


# ============================================================
# Export Endpoints
# ============================================================

@app.get("/api/sets/{set_id}/export/excel")
def export_excel(set_id: int, db: Session = Depends(get_db)):
    """Export set to Excel file."""
    from fastapi.responses import FileResponse
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import tempfile
    import os
    
    card_set = db.query(CardSet).filter(CardSet.id == set_id).first()
    if not card_set:
        raise HTTPException(404, "Set not found")
    
    cards = db.query(Card).filter(Card.set_id == set_id).all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = card_set.name
    
    headers = ["Card #", "Player", "Team", "RC/SP", "Insert Type", "Parallel", "Qty"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'), bottom=Side(style='thin'))
    
    for col, h in enumerate(headers, 1):
        cell = ws.cell(1, col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin
    
    for i, card in enumerate(cards, 2):
        ws.cell(i, 1, value=card.card_number).border = thin
        ws.cell(i, 2, value=card.player).border = thin
        ws.cell(i, 3, value=card.team).border = thin
        ws.cell(i, 4, value=card.rc_sp).border = thin
        ws.cell(i, 5, value=card.insert_type).border = thin
        ws.cell(i, 6, value=card.parallel).border = thin
        ws.cell(i, 7, value=card.qty if card.qty > 0 else None).border = thin
    
    # Auto-width columns
    for col in range(1, 8):
        ws.column_dimensions[chr(64 + col)].width = 18
    
    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    wb.save(tmp.name)
    tmp.close()
    
    return FileResponse(
        tmp.name,
        filename=f"{card_set.name}.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.get("/api/sets/{set_id}/export/csv")
def export_csv(set_id: int, db: Session = Depends(get_db)):
    """Export set to CSV (for eBay variation listings)."""
    from fastapi.responses import StreamingResponse
    import csv
    import io
    
    card_set = db.query(CardSet).filter(CardSet.id == set_id).first()
    if not card_set:
        raise HTTPException(404, "Set not found")
    
    cards = db.query(Card).filter(Card.set_id == set_id, Card.qty > 0).all()
    
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Card #", "Player", "Team", "RC/SP", "Insert Type", "Parallel", "Qty"])
    
    for card in cards:
        writer.writerow([
            card.card_number, card.player, card.team,
            card.rc_sp, card.insert_type, card.parallel, card.qty
        ])
    
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={card_set.name}.csv"}
    )


# ============================================================
# WebSocket for Real-Time Voice
# ============================================================

@app.websocket("/ws/voice")
async def voice_websocket(websocket: WebSocket):
    """
    WebSocket endpoint for real-time voice processing.
    
    Client sends: audio chunks (binary) or JSON commands
    Server sends: JSON with parsed numbers and status updates
    
    Commands:
        {"action": "start", "set_id": 1, "insert_type": "Base"}
        {"action": "stop"}
        {"action": "clear"}
        {"action": "parse", "text": "42 55 103 times 2"}
    """
    await websocket.accept()
    
    try:
        while True:
            data = await websocket.receive_text()
            msg = json.loads(data)
            action = msg.get("action", "")
            
            if action == "parse":
                # Parse text input (from frontend text box or speech API)
                text = msg.get("text", "")
                numbers = parse_spoken_numbers(text)
                await websocket.send_json({
                    "type": "numbers",
                    "numbers": numbers,
                    "counts": count_cards(numbers),
                    "output": format_output(numbers),
                    "raw": text,
                })
            
            elif action == "ping":
                await websocket.send_json({"type": "pong"})
    
    except WebSocketDisconnect:
        pass


# ============================================================
# Health Check
# ============================================================

@app.get("/api/health")
def health():
    return {"status": "ok", "version": "0.2.0", "db_path": DB_PATH, "db_dir": DB_DIR}
